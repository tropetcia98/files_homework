from zipfile import ZipFile
from set_path import *
from pypdf import PdfReader
from openpyxl import load_workbook
import csv


def test_pdf():
    with ZipFile(ZIP_FILE) as zf:
        with zf.open('english_for_everyone_level_1_course_book_beginner.pdf', 'r') as pdf_file:
            reader = PdfReader(pdf_file)
            text = reader.pages[1].extract_text()
            assert 'Rachel Harding has a background in English-language teaching' in text


def test_xlsx():
    with ZipFile(ZIP_FILE) as zf:
        with zf.open('file_example_XLSX_50.xlsx', 'r') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            cell_value = sheet.cell(row=3, column=2).value
            assert 'Mara' in cell_value


def test_csv():
    with ZipFile(ZIP_FILE) as zf:
        with zf.open('industry.csv', 'r') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csv_reader = list(csv.reader(content.splitlines()))
            twenty_row = csv_reader[20]
            assert 'Human Resources' in twenty_row
